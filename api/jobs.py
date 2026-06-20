import csv
import hashlib
import io
import json
import os
import re
import ssl
import urllib.parse
import urllib.request
import uuid
from http.server import BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
from datetime import datetime, date

import pg8000.dbapi

# Auth — prefer APP_PASSWORD from env, but fall back to Ratul's chosen code.
PASSWORD = os.environ.get("APP_PASSWORD", "ratul2026")
VALID_TOKEN = ("jt-" + hashlib.sha256(("v1:" + PASSWORD).encode()).hexdigest()[:40]) if PASSWORD else ""

# Database (Neon Postgres). DATABASE_URL = postgresql://user:pass@host/db?sslmode=require
DATABASE_URL = os.environ.get("DATABASE_URL", "")
_conn = None

# Columns owned by the app (search_blob is a generated column, never written directly).
JOB_COLUMNS = [
    "id", "company", "role", "tier", "category", "url", "careers_url",
    "salary_range", "location", "status", "notes", "description", "source",
    "date_added", "date_applied", "priority", "deadline",
]
DATE_FIELDS = {"date_added", "date_applied", "deadline"}


def _parse_dsn(url):
    p = urlparse(url)
    return dict(
        user=p.username,
        password=p.password,
        host=p.hostname,
        port=p.port or 5432,
        database=(p.path or "/").lstrip("/") or "neondb",
    )


def get_conn():
    """Cached connection, revalidated per request (Neon pooled endpoint)."""
    global _conn
    if _conn is not None:
        try:
            c = _conn.cursor()
            c.execute("SELECT 1")
            c.close()
            return _conn
        except Exception:
            try:
                _conn.close()
            except Exception:
                pass
            _conn = None
    ctx = ssl.create_default_context()
    _conn = pg8000.dbapi.connect(ssl_context=ctx, **_parse_dsn(DATABASE_URL))
    _conn.autocommit = True
    return _conn


def _iso(v):
    if isinstance(v, (date, datetime)):
        return v.isoformat()[:10]
    return v


def _date_or_none(v):
    v = v.strip() if isinstance(v, str) else v
    return v if v else None


def row_to_job(row):
    job = {}
    for k, v in zip(JOB_COLUMNS, row):
        job[k] = _iso(v) if k in DATE_FIELDS else v
    return job
ARCHIVE_STATUSES = {"expired", "not_interested", "rejected", "withdrawn"}
AUTO_KEYWORDS = [
    'automation', 'zapier', 'airtable', 'no-code', 'nocode', 'integration', 'crm', 'api',
    'make.com', 'n8n', 'smartlead', 'clay', 'revops', 'revenue operations', 'sales ops',
    'gtm ops', 'workflow'
]
EARLY_GOOD = [
    'early career', 'early careers', 'new grad', 'new graduate', 'graduate program', 'campus',
    'intern', 'entry level', 'analyst', 'associate', 'coordinator', 'specialist', 'trainee', 'rotational'
]
EARLY_BAD = [
    '5+ year', '5 years', '6+ year', '7+ year', '8+ year', '10+ year', 'senior manager',
    'director', 'vice president', 'vp ', 'principal', 'partner', 'cpa required', 'isda', 'trioptima'
]
CFA_KEYWORDS = [
    'cfa', 'investment analy', 'portfolio', 'asset manage', 'wealth manage', 'fund account',
    'fund admin', 'valuation', 'financial analy', 'risk analy', 'model risk', 'credit risk',
    'fixed income', 'equity research', 'securities', 'capital market', 'corporate finance',
    'fp&a', 'financial plan', 'actuari', 'insurance analy', 'reits', 'real estate invest', 'cfa-qualifying'
]
FINANCE_FALL_GOOD = [
    'intern', 'internship', 'fall internship', 'fall 2026', 'autumn 2026', 'campus', 'new grad',
    'new graduate', 'analyst', 'associate', 'equity analyst', 'equity research', 'research analyst',
    'investment analyst', 'quant', 'quantitative', 'trading analyst', 'markets analyst', 'capital markets',
    'portfolio analyst', 'credit analyst', 'risk analyst', 'financial analyst'
]
FINANCE_FALL_DOMAIN = [
    'private equity', 'investment banking', 'equity research', 'asset management', 'wealth', 'pension',
    'capital markets', 'trading', 'quant', 'valuation', 'credit', 'portfolio', 'securities', 'investments',
    'corporate finance', 'financial modeling', 'fp&a', 'commercial finance', 'treasury'
]
FINANCE_FALL_BAD = [
    'senior ', 'director', 'vice president', 'vp ', 'principal', 'head of', 'chief ',
    '5+ year', '6+ year', '7+ year', '8+ year', '10+ year', 'manager,', 'manager -', 'manager –'
]
MICRO_MATURE_MARKERS = [
    'micro mature target', 'micro-mature target', 'micro mature', '1-10 employees', '1–10 employees',
    '5+ years old', 'at least 5 years old', '$5m+ revenue', '5m+ revenue', '5m revenue'
]
BOUTIQUE_HINTS = [
    'partners', 'capital', 'advisors', 'advisory', 'securities', 'wealth', 'ventures', 'equity', 'holdings'
]
TOP_TIER_FIRMS = [
    'rbc', 'td', 'bmo', 'scotiabank', 'cibc', 'national bank', 'brookfield', 'onex', 'otp',
    'ontario teachers', 'hoopp', 'cpp investments', 'cpb investments', 'omers', 'kpmg', 'deloitte',
    'pwc', 'ey', 'mckinsey', 'bcg', 'bain', 'goldman sachs', 'jpmorgan', 'morgan stanley', 'blackrock'
]
P_MAP = {"high": 0, "medium": 1, "low": 2}


_SELECT = "SELECT " + ", ".join(JOB_COLUMNS) + " FROM jobs"


def load_all_jobs():
    cur = get_conn().cursor()
    cur.execute(_SELECT)
    rows = cur.fetchall()
    cur.close()
    return {"jobs": [row_to_job(r) for r in rows]}


def get_job(job_id):
    cur = get_conn().cursor()
    cur.execute(_SELECT + " WHERE id = %s", (job_id,))
    row = cur.fetchone()
    cur.close()
    return row_to_job(row) if row else None


def is_rahat_job(job):
    """Rahat's auto-found Canada jobs are tagged source='rahat-*'. This is the
    partition key that keeps his board separate from Ratul's tabs — no schema
    change needed."""
    return (job.get("source") or "").startswith("rahat")


def build_meta(jobs):
    # Meta drives Ratul's stat pills + category dropdown — exclude Rahat's jobs.
    jobs = [j for j in jobs if not is_rahat_job(j)]
    counts = {}
    categories = {}
    for job in jobs:
        status = job.get("status", "unknown")
        counts[status] = counts.get(status, 0) + 1
        category = (job.get("category") or "Other").strip()
        categories[category] = categories.get(category, 0) + 1
    active = [j for j in jobs if j.get("status") not in ARCHIVE_STATUSES]
    archived = [j for j in jobs if j.get("status") in ARCHIVE_STATUSES]
    return {
        "total": len(jobs),
        "active": len(active),
        "archived": len(archived),
        "counts": counts,
        "categories": dict(sorted(categories.items(), key=lambda kv: (-kv[1], kv[0].lower()))),
    }


def insert_job(job):
    cols = [c for c in JOB_COLUMNS]
    vals = [_date_or_none(job.get(c)) if c in DATE_FIELDS else job.get(c) for c in cols]
    sql = (
        "INSERT INTO jobs (" + ", ".join(cols) + ") VALUES ("
        + ", ".join(["%s"] * len(cols)) + ")"
    )
    cur = get_conn().cursor()
    cur.execute(sql, vals)
    cur.close()


def update_job(job_id, fields):
    if not fields:
        return
    sets, vals = [], []
    for k, v in fields.items():
        sets.append(f"{k} = %s")
        vals.append(_date_or_none(v) if k in DATE_FIELDS else v)
    sets.append("updated_at = now()")
    vals.append(job_id)
    cur = get_conn().cursor()
    cur.execute("UPDATE jobs SET " + ", ".join(sets) + " WHERE id = %s", vals)
    cur.close()


def delete_job(job_id):
    cur = get_conn().cursor()
    cur.execute("DELETE FROM jobs WHERE id = %s", (job_id,))
    rowcount = cur.rowcount
    cur.close()
    return rowcount > 0


def cors_headers():
    return {
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "GET, POST, PATCH, DELETE, OPTIONS",
        "Access-Control-Allow-Headers": "Content-Type, X-Auth-Token",
        "Content-Type": "application/json"
    }


def _s(value):
    return "" if value is None else str(value)


def text_blob(job):
    return ((
        _s(job.get("role")) + " " +
        _s(job.get("category")) + " " +
        _s(job.get("company")) + " " +
        _s(job.get("notes")) + " " +
        _s(job.get("description")) + " " +
        _s(job.get("location"))
    )).lower()


def infer_firm_label(job):
    text = text_blob(job)
    company = _s(job.get('company')).lower()
    category = _s(job.get('category')).lower()
    if any(name in company for name in TOP_TIER_FIRMS):
        return 'Top Tier'
    if 'boutique' in text:
        return 'Boutique'
    if any(hint in company for hint in BOUTIQUE_HINTS):
        return 'Boutique'
    if any(k in text for k in ['pension', 'teachers', 'hoopp', 'cpp investments', 'cpb investments', 'omers']):
        return 'Institutional'
    if any(k in text for k in ['bank', 'banking', 'capital markets', 'securities']):
        return 'Large Firm'
    if any(k in category for k in ['private equity', 'investment banking', 'venture capital', 'capital markets', 'asset', 'wealth']):
        return 'Finance Firm'
    return 'Other'


def is_fall_finance_target(job):
    text = text_blob(job)
    if any(bad in text for bad in FINANCE_FALL_BAD):
        return False
    has_role_signal = any(good in text for good in FINANCE_FALL_GOOD)
    has_domain_signal = any(dom in text for dom in FINANCE_FALL_DOMAIN)
    if not (has_role_signal and has_domain_signal):
        return False
    status = job.get('status')
    if status in {'expired', 'withdrawn', 'rejected', 'not_interested'}:
        return False
    return True


def score_match(job):
    text = text_blob(job)
    lanes = {
        "Integrations": {
            "keywords": ['airtable', 'zapier', 'crm', 'hubspot', 'salesforce', 'integration', 'integrations', 'automation', 'workflow', 'api', 'implementation', 'onboarding', 'revops', 'revenue operations', 'sales ops', 'ops', 'support', 'technical support', 'customer success', 'solutions'],
            "titleBoost": ['specialist', 'coordinator', 'analyst', 'implementation', 'operations'],
        },
        "FPA": {
            "keywords": ['financial analyst', 'fp&a', 'forecasting', 'budgeting', 'cost analysis', 'pricing', 'p&l', 'variance', 'commercial finance', 'operational finance', 'reporting', 'excel', 'power bi', 'erp'],
            "titleBoost": ['analyst', 'finance', 'operations'],
        },
        "Data Analyst": {
            "keywords": ['data analyst', 'business analyst', 'analytics', 'dashboard', 'sql', 'python', 'tableau', 'power bi', 'etl', 'data pipeline', 'reporting', 'metrics', 'insights', 'product analytics'],
            "titleBoost": ['analyst', 'analytics', 'data'],
        },
        "PE": {
            "keywords": ['investment analyst', 'portfolio analytics', 'institutional investing', 'manager research', 'asset allocation', 'equity research', 'valuation', 'capital markets', 'cfa', 'investment consulting', 'wealth', 'pension', 'transaction advisory', 'corporate finance', 'underwriting', 'due diligence', 'acquisition', 'portfolio company', 'private markets', 'private capital', 'm&a', 'deals'],
            "titleBoost": ['analyst', 'associate', 'investment', 'financial analyst'],
        },
        "Consulting": {
            "keywords": ['consulting', 'strategy', 'business operations', 'project management', 'stakeholder', 'process improvement', 'cross-functional', 'operations analyst'],
            "titleBoost": ['analyst', 'associate', 'consulting'],
        }
    }
    negatives_hard = ['10+ year', '8+ year', '7+ year', '6+ year', '5-8 years', '5 to 8 years', 'director', 'vice president', 'vp ', 'principal', 'partner', 'cpa required', 'isda', 'csa', 'trioptima', 'numerix', 'markit', 'aladdin']
    negatives_medium = ['4+ year', '4 years', '5+ year', '5 years', 'senior manager', 'senior ', 'lead ', 'manager', 'derivatives middle office', 'quant developer', 'insurance product forms']
    context_bonus = ['entry level', 'new grad', 'new graduate', 'early career', 'analyst', 'associate', 'coordinator', 'specialist', 'toronto', 'waterloo', 'remote', 'hybrid', 'canada', 'startup', 'saas', 'fintech']
    interview_odds_bonus = ['airtable', 'zapier', 'clay', 'greenhouse', 'business systems', 'technical solutions', 'implementation', 'enablement', 'customer success', 'crm', 'revops', 'revenue operations', 'sales operations', 'workflow', 'api', 'webhook', 'dashboard', 'reporting', 'due diligence', 'underwriting', 'acquisition', 'screening', 'valuation', 'portfolio company', 'corporate finance', 'transaction advisory', 'deals', 'm&a', 'wealth management', 'asset management', 'private markets', 'private capital']
    interview_odds_penalty = ['staffing', 'recruiter', 'recruiting', 'contract', '12 month contract', '6 month contract', 'senior analyst', 'associate director', 'trading', 'quant', 'supervisory analyst']
    best_resume = 'General'
    best_score = -999
    reasons = []
    for resume, cfg in lanes.items():
        score = 0
        local_reasons = []
        for kw in cfg['keywords']:
            if kw in text:
                score += 4
                if len(local_reasons) < 3:
                    local_reasons.append(kw)
        for kw in cfg['titleBoost']:
            if kw in text:
                score += 2
        for kw in context_bonus:
            if kw in text:
                score += 1
        for kw in interview_odds_bonus:
            if kw in text:
                score += 2
        for kw in negatives_hard:
            if kw in text:
                score -= 8
        for kw in negatives_medium:
            if kw in text:
                score -= 4
        for kw in interview_odds_penalty:
            if kw in text:
                score -= 3
        if ((('asset management' in text) or ('private equity' in text) or ('equity research' in text) or ('portfolio' in text))
                and resume != 'PE' and not (('new grad' in text) or ('entry level' in text) or ('intern' in text))):
            score -= 3
        if any(k in text for k in ['startup', 'saas', 'fintech', 'healthtech', 'software', 'scale-up', 'scaleup', 'mid-sized', 'mid sized', 'small company', 'growing company', 'boutique', 'independent firm', 'wealth management firm', 'asset management firm', 'advisory firm', 'holding company', 'family office']):
            score += 3
        if 'toronto' in text:
            score += 2
        if any(k in text for k in ['posted today', 'posted 1 hour', 'posted 2 hours', 'posted 3 hours', 'posted 4 hours', 'posted 5 hours', 'posted 6 hours', 'posted 7 hours', 'posted 8 hours', 'posted 9 hours', 'posted 10 hours', 'posted 11 hours', 'posted 12 hours', 'posted 13 hours', 'posted 14 hours', 'posted 15 hours', 'posted 16 hours', 'posted 17 hours', 'posted 18 hours', 'posted 19 hours', 'posted 20 hours', 'posted 21 hours', 'posted 22 hours', 'posted 23 hours', 'posted within 24h', 'fresh linkedin posting']):
            score += 2
        status = job.get('status')
        if status == 'not_applied':
            score += 2
        elif status == 'saved':
            score += 1
        elif status in {'expired', 'withdrawn', 'rejected', 'not_interested'}:
            score -= 12
        if score > best_score:
            best_score = score
            best_resume = resume
            reasons = local_reasons
    final_score = max(best_score, 0)
    if final_score >= 22:
        chance = 'High Chance'
    elif final_score >= 14:
        chance = 'Strong Fit'
    elif final_score >= 8:
        chance = 'Reach'
    else:
        chance = 'Low ROI'
    return {
        'score': final_score,
        'bestResume': best_resume,
        'chance': chance,
        'reasons': reasons,
    }


def apply_filters(jobs, qs):
    status_filter = qs.get('status', [''])[0]
    category = qs.get('category', [''])[0]
    q = qs.get('q', [''])[0].strip().lower()
    tab = qs.get('tab', ['all'])[0]
    sort = qs.get('sort', ['date_added'])[0]
    sort_dir = qs.get('dir', [''])[0]
    qmode = qs.get('qmode', ['basic'])[0]
    exclude_terms = [t.strip().lower() for t in qs.get('exclude', [''])[0].split(',') if t.strip()]
    categories = [c for c in qs.get('categories', [''])[0].split(',') if c]
    statuses = [s for s in qs.get('statuses', [''])[0].split(',') if s]

    # Multi-term search: every whitespace-separated token must match (AND).
    q_terms = [t for t in q.split() if t]

    filtered = []
    for job in jobs:
        text = text_blob(job)
        rahat = is_rahat_job(job)
        # Partition: the 'rahat' tab shows only Rahat's jobs; every other tab
        # (incl. 'all', 'search', 'bestmatch') excludes them.
        if tab == 'rahat':
            if not rahat:
                continue
        elif rahat:
            continue
        if statuses:
            if job.get('status') not in statuses:
                continue
        elif status_filter and job.get('status') != status_filter:
            continue
        if categories:
            if job.get('category') not in categories:
                continue
        elif category and job.get('category') != category:
            continue
        if q_terms:
            # Quick search (curated tabs) spans company/role/category/location;
            # the Search tab's full mode spans the whole blob (notes/desc too).
            haystack = text if qmode == 'full' else (
                f"{job.get('company', '')} {job.get('role', '')} "
                f"{job.get('category', '')} {job.get('location', '')}"
            ).lower()
            if not all(term in haystack for term in q_terms):
                continue
        if exclude_terms and any(term in text for term in exclude_terms):
            continue
        if tab == 'automation' and not any(kw in text for kw in AUTO_KEYWORDS):
            continue
        if tab == 'earlycareers':
            if not any(kw in text for kw in EARLY_GOOD):
                continue
            if any(kw in text for kw in EARLY_BAD):
                continue
        if tab == 'cfa' and not any(kw in text for kw in CFA_KEYWORDS):
            continue
        if tab == 'fallfinance' and not is_fall_finance_target(job):
            continue
        if tab == 'applied' and job.get('status') != 'applied':
            continue
        if tab == 'interviews' and job.get('status') != 'interview':
            continue

        job = dict(job)
        job['_firmLabel'] = infer_firm_label(job)

        if tab == 'bestmatch':
            match = score_match(job)
            if match['score'] < 5:
                continue
            job['_matchScore'] = match['score']
            job['_bestResume'] = match['bestResume']
            job['_chance'] = match['chance']
            job['_matchReasons'] = match['reasons']
        elif tab == 'fallfinance':
            match = score_match(job)
            job['_matchScore'] = match['score']
            job['_bestResume'] = match['bestResume']
            job['_chance'] = match['chance']
            job['_matchReasons'] = match['reasons']
        filtered.append(job)

    SORT_KEYS = {
        'company': lambda j: (j.get('company') or '').lower(),
        'role': lambda j: (j.get('role') or '').lower(),
        'category': lambda j: (j.get('category') or '').lower(),
        'status': lambda j: (j.get('status') or '').lower(),
        'location': lambda j: (j.get('location') or '').lower(),
        'priority': lambda j: P_MAP.get(j.get('priority', 'medium'), 1),
        'date_added': lambda j: j.get('date_added') or '',
        'date_applied': lambda j: j.get('date_applied') or '',
        'deadline': lambda j: j.get('deadline') or '',
    }
    if tab == 'search' and sort in SORT_KEYS:
        reverse = (sort_dir == 'desc') if sort_dir else (sort in ('date_added', 'date_applied', 'deadline'))
        filtered.sort(key=SORT_KEYS[sort], reverse=reverse)
    elif sort == 'priority':
        filtered.sort(key=lambda j: (P_MAP.get(j.get('priority', 'medium'), 1), -(j.get('_matchScore') or 0)))
    elif sort == 'company':
        filtered.sort(key=lambda j: j.get('company', '').lower())
    elif tab == 'bestmatch':
        filtered.sort(key=lambda j: j.get('_matchScore', 0), reverse=True)
    elif tab == 'fallfinance':
        firm_rank = {'Top Tier': 0, 'Large Firm': 1, 'Institutional': 2, 'Boutique': 3, 'Finance Firm': 4, 'Other': 5}
        filtered.sort(key=lambda j: (firm_rank.get(j.get('_firmLabel'), 9), -(j.get('_matchScore') or 0), j.get('company', '').lower()))
    elif tab == 'micromature':
        firm_rank = {'Boutique': 0, 'Finance Firm': 1, 'Other': 2, 'Large Firm': 3, 'Institutional': 4, 'Top Tier': 5}
        filtered.sort(key=lambda j: (firm_rank.get(j.get('_firmLabel'), 9), P_MAP.get(j.get('priority', 'medium'), 1), j.get('company', '').lower()))
    else:
        reverse = (sort_dir != 'asc')
        filtered.sort(key=lambda j: j.get('date_added') or '', reverse=reverse)
    return filtered


def summarize_job(job):
    keep = [
        'id', 'company', 'role', 'tier', 'category', 'url', 'careers_url', 'salary_range', 'location',
        'status', 'notes', 'date_added', 'date_applied', 'priority', 'deadline',
        '_matchScore', '_bestResume', '_chance', '_matchReasons', '_firmLabel'
    ]
    out = {k: job.get(k) for k in keep if k in job}
    if out.get('notes') and len(out['notes']) > 240:
        out['notes_preview'] = out['notes'][:240] + '…'
    return out


# ===========================================================================
# Rahat (Canada software/tech) auto-ingest — a daily Vercel cron hits /api/cron.
# Jobs are inserted with source='rahat-*' so they live only in the Rahat tab.
# Profile: Mohammad Rahat Hasan — Java/SQL/JavaScript/Angular/Power Platform/
# SSIS/AI-LLM; junior–mid (~2 yrs, technical-delivery consultant). Targets CA.
# Source: Adzuna CA (free key, optional) primary + Remotive keyless fallback.
# ===========================================================================
RAHAT_QUERIES = [
    "software developer",
    "java developer",
    "full stack developer",
    "power platform developer",
    "technical consultant",
    "data analyst",
]
RAHAT_SKILLS = [
    'java', 'sql', 'mysql', 'sql server', 'javascript', 'typescript', 'angular',
    'react', 'html', 'css', 'python', 'power platform', 'powerapps', 'power apps',
    'power automate', 'power bi', 'ssis', 'rest api', 'api', 'integration',
    'automation', 'full stack', 'fullstack', 'full-stack', 'software engineer',
    'software developer', 'developer', 'data analyst', 'business analyst',
    'technical consultant', 'low-code', 'low code', 'agile', 'scrum', 'sdlc',
    'llm', 'ai', 'machine learning', 'application developer', 'web developer',
    'frontend', 'backend', 'front-end', 'back-end', '.net', 'c#',
]
RAHAT_NEG = [
    'senior ', 'staff ', 'principal', 'lead ', 'manager', 'director', 'head of',
    'vice president', 'vp ', 'architect', '5+ year', '6+ year', '7+ year',
    '8+ year', '10+ year', 'security clearance', 'secret clearance', 'phd',
]
RAHAT_MAX_INSERT = 25  # bound per-run work so the serverless function stays well under timeout


def _http_json(url, timeout=6, headers=None):
    req = urllib.request.Request(url, headers=headers or {"User-Agent": "job-tracker-cron/1.0"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8", "replace"))


def _norm_url(u):
    return (u or "").split("?")[0].rstrip("/").lower()


def rahat_match(role, body):
    text = (str(role) + " " + str(body)).lower()
    hits = []
    for s in RAHAT_SKILLS:
        # Word-boundary match (non-alphanumeric edges) so 'ssis' doesn't match
        # 'assistant', 'api' doesn't match 'capital', 'ai' doesn't match 'available'.
        if re.search(r"(?<![a-z0-9])" + re.escape(s) + r"(?![a-z0-9])", text):
            hits.append(s)
    neg = any(n in text for n in RAHAT_NEG)
    return sorted(set(hits)), neg


def _adzuna_salary(r):
    lo, hi = r.get("salary_min"), r.get("salary_max")
    if lo and hi:
        return f"${int(lo):,}–${int(hi):,}"
    return None


def fetch_adzuna_ca():
    app_id = os.environ.get("ADZUNA_APP_ID", "")
    app_key = os.environ.get("ADZUNA_APP_KEY", "")
    if not (app_id and app_key):
        return []
    out = []
    for what in RAHAT_QUERIES:
        params = urllib.parse.urlencode({
            "app_id": app_id, "app_key": app_key, "what": what,
            "results_per_page": 20, "max_days_old": 7, "sort_by": "date",
            "content-type": "application/json",
        })
        url = f"https://api.adzuna.com/v1/api/jobs/ca/search/1?{params}"
        try:
            data = _http_json(url)
        except Exception as e:
            print(f"adzuna fetch error ({what}): {e}")
            continue
        for r in data.get("results", []):
            out.append({
                "company": (r.get("company") or {}).get("display_name") or "Unknown",
                "role": r.get("title") or "",
                "url": r.get("redirect_url") or "",
                "location": (r.get("location") or {}).get("display_name") or "Canada",
                "salary_range": _adzuna_salary(r),
                "description": r.get("description") or "",
                "source": "rahat-adzuna",
            })
    return out


def _strip_html(s):
    s = re.sub(r"<[^>]+>", " ", s or "")
    s = s.replace("&amp;", "&").replace("&nbsp;", " ").replace("&#39;", "'")
    return re.sub(r"\s+", " ", s).strip()


# Remote-eligible location tokens a Canada-based applicant can take.
_CA_OK = ["worldwide", "anywhere", "canada", "north america", "americas", "usa, canada", "usa/canada"]


def fetch_remotive_ca():
    """Keyless fallback — Remotive remote jobs filtered to Canada-eligible
    locations (Canada / North America / Worldwide). Better location metadata
    than Jobicy, which buckets by region and returns ~nothing for Canada."""
    out = []
    endpoints = [
        "https://remotive.com/api/remote-jobs?category=software-dev&limit=100",
        "https://remotive.com/api/remote-jobs?search=data+analyst&limit=50",
    ]
    for url in endpoints:
        try:
            data = _http_json(url, timeout=8)
        except Exception as e:
            print(f"remotive fetch error: {e}")
            continue
        for r in data.get("jobs", []):
            loc = (r.get("candidate_required_location") or "").lower()
            if not any(tok in loc for tok in _CA_OK):
                continue
            out.append({
                "company": r.get("company_name") or "Unknown",
                "role": r.get("title") or "",
                "url": r.get("url") or "",
                "location": (r.get("candidate_required_location") or "Remote") + " · Remote",
                "salary_range": (r.get("salary") or "").strip() or None,
                "description": _strip_html(r.get("description"))[:800],
                "source": "rahat-remotive",
            })
    return out


def ingest_rahat():
    existing = load_all_jobs().get("jobs", [])
    seen_urls, seen_pairs = set(), set()
    for j in existing:
        if is_rahat_job(j):
            if j.get("url"):
                seen_urls.add(_norm_url(j["url"]))
            seen_pairs.add(((j.get("company") or "").lower().strip(),
                            (j.get("role") or "").lower().strip()))

    candidates = []
    for fetch in (fetch_adzuna_ca, fetch_remotive_ca):
        try:
            candidates += fetch()
        except Exception as e:
            print(f"rahat fetch error ({fetch.__name__}): {e}")

    added, dup, unfit = 0, 0, 0
    today = datetime.now().strftime("%Y-%m-%d")
    for c in candidates:
        if added >= RAHAT_MAX_INSERT:
            break
        company, role = (c.get("company") or "").strip(), (c.get("role") or "").strip()
        if not company or not role:
            continue
        nu = _norm_url(c.get("url"))
        pair = (company.lower(), role.lower())
        if (nu and nu in seen_urls) or pair in seen_pairs:
            dup += 1
            continue
        hits, neg = rahat_match(role, c.get("description", ""))
        if neg or not hits:
            unfit += 1
            continue
        seen_urls.add(nu)
        seen_pairs.add(pair)
        priority = "high" if len(hits) >= 4 else ("medium" if len(hits) >= 2 else "low")
        src = (c.get("source") or "rahat-auto").replace("rahat-", "")
        job = {
            "id": str(uuid.uuid4()),
            "company": company,
            "role": role,
            "tier": 4,
            "category": "Software / Tech",
            "url": c.get("url") or "",
            "careers_url": "",
            "salary_range": c.get("salary_range"),
            "location": c.get("location") or "Canada",
            "status": "not_applied",
            "notes": f"Auto-found for Rahat via {src}. Skills matched: {', '.join(hits[:8])}.",
            "description": (c.get("description") or "")[:1500],
            "source": c.get("source") or "rahat-auto",
            "date_added": today,
            "date_applied": None,
            "priority": priority,
            "deadline": None,
        }
        try:
            insert_job(job)
            added += 1
        except Exception as e:
            print(f"rahat insert error: {e}")
    return {"added": added, "skipped_duplicate": dup, "skipped_unfit": unfit,
            "candidates": len(candidates)}


# --- Export (CSV / Excel) -- reflects the exact filtered set, no pagination ---
EXPORT_COLUMNS = [
    ("company", "Company"), ("role", "Role"), ("category", "Category"),
    ("status", "Status"), ("priority", "Priority"), ("location", "Location"),
    ("salary_range", "Salary"), ("url", "URL"), ("date_added", "Date Added"),
    ("date_applied", "Date Applied"), ("deadline", "Deadline"),
    ("source", "Source"), ("notes", "Notes"),
]


def _cell(v):
    return "" if v is None else str(v)


def export_csv(jobs):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([label for _, label in EXPORT_COLUMNS])
    for j in jobs:
        w.writerow([_cell(j.get(k)) for k, _ in EXPORT_COLUMNS])
    # Prepend UTF-8 BOM so Excel opens accented text + columns cleanly.
    return ("﻿" + buf.getvalue()).encode("utf-8")


def export_xlsx(jobs):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    headers = [label for _, label in EXPORT_COLUMNS]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="111827")
    font = Font(bold=True, color="FFFFFF")
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill, c.font = fill, font
        c.alignment = Alignment(vertical="center")
    for j in jobs:
        ws.append([_cell(j.get(k)) for k, _ in EXPORT_COLUMNS])
    widths = [22, 40, 18, 13, 10, 24, 16, 42, 12, 12, 12, 14, 60]
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = ws.dimensions
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


class handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass

    def send_json(self, status, data):
        body = json.dumps(data).encode()
        self.send_response(status)
        for k, v in cors_headers().items():
            self.send_header(k, v)
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_bytes(self, status, data, content_type, filename):
        self.send_response(status)
        for k, v in cors_headers().items():
            if k == 'Content-Type':
                continue
            self.send_header(k, v)
        self.send_header('Content-Type', content_type)
        self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
        self.send_header('Content-Length', str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def check_auth(self):
        token = self.headers.get('X-Auth-Token', '')
        return bool(VALID_TOKEN) and token == VALID_TOKEN

    def do_OPTIONS(self):
        self.send_response(204)
        for k, v in cors_headers().items():
            self.send_header(k, v)
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        path = parsed.path.rstrip('/')

        if path == '/api/auth':
            pw = qs.get('password', [''])[0]
            if PASSWORD and pw == PASSWORD:
                return self.send_json(200, {'token': VALID_TOKEN, 'ok': True})
            return self.send_json(401, {'error': 'Invalid password'})

        if path == '/api/cron':
            # Vercel attaches `Authorization: Bearer $CRON_SECRET` if the env var
            # is set. If set, require it (also accept ?key= for manual triggers);
            # if unset, run open so it works before you configure the secret.
            secret = os.environ.get('CRON_SECRET', '')
            if secret:
                auth = self.headers.get('Authorization', '')
                if auth != f'Bearer {secret}' and qs.get('key', [''])[0] != secret:
                    return self.send_json(401, {'error': 'Unauthorized'})
            try:
                result = ingest_rahat()
            except Exception as e:
                print(f"cron error: {e}")
                return self.send_json(500, {'error': 'cron failed'})
            return self.send_json(200, {'ok': True, **result})

        if path == '/api/export':
            try:
                jobs = load_all_jobs().get('jobs', [])
            except Exception as e:
                print(f"DB read error: {e}")
                return self.send_json(500, {'error': 'Database error'})
            filtered = apply_filters(jobs, qs)
            fmt = qs.get('format', ['csv'])[0].lower()
            tab = qs.get('tab', ['all'])[0]
            stamp = datetime.now().strftime('%Y%m%d')
            if fmt in ('xlsx', 'excel'):
                try:
                    data = export_xlsx(filtered)
                    return self.send_bytes(
                        200, data,
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        f'jobs_{tab}_{stamp}.xlsx')
                except Exception as e:
                    print(f"xlsx export error (falling back to csv): {e}")
            data = export_csv(filtered)
            return self.send_bytes(200, data, 'text/csv; charset=utf-8',
                                   f'jobs_{tab}_{stamp}.csv')

        if path == '/api/jobs':
            job_id = qs.get('id', [''])[0]
            try:
                if job_id:
                    job = get_job(job_id)
                    if not job:
                        return self.send_json(404, {'error': 'Job not found'})
                    return self.send_json(200, job)
                jobs = load_all_jobs().get('jobs', [])
            except Exception as e:
                print(f"DB read error: {e}")
                return self.send_json(500, {'error': 'Database error'})

            filtered = apply_filters(jobs, qs)
            total_filtered = len(filtered)
            try:
                page = max(int(qs.get('page', ['1'])[0]), 1)
            except Exception:
                page = 1
            try:
                page_size = int(qs.get('page_size', ['40'])[0])
            except Exception:
                page_size = 40
            page_size = max(1, min(page_size, 100))
            start = (page - 1) * page_size
            end = start + page_size
            paged = filtered[start:end]
            slim = qs.get('detail', ['summary'])[0] != 'full'
            jobs_out = [summarize_job(j) for j in paged] if slim else paged
            return self.send_json(200, {
                'jobs': jobs_out,
                'meta': build_meta(jobs),
                'page': page,
                'page_size': page_size,
                'total_filtered': total_filtered,
                'has_more': end < total_filtered,
            })

        self.send_json(404, {'error': 'Not found'})

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path.rstrip('/')
        if path != '/api/jobs':
            return self.send_json(404, {'error': 'Not found'})
        length = int(self.headers.get('Content-Length', 0))
        body = json.loads(self.rfile.read(length).decode())
        try:
            tier = int(body.get('tier', 4))
        except Exception:
            tier = 4
        new_job = {
            'id': str(uuid.uuid4()),
            'company': body.get('company', ''),
            'role': body.get('role', ''),
            'tier': tier,
            'category': body.get('category', 'Business Operations'),
            'url': body.get('url', ''),
            'careers_url': body.get('careers_url', ''),
            'salary_range': body.get('salary_range', ''),
            'location': body.get('location', 'Toronto'),
            'status': body.get('status', 'not_applied'),
            'notes': body.get('notes', ''),
            'description': body.get('description', ''),
            'source': body.get('source', 'manual'),
            'date_added': datetime.now().strftime('%Y-%m-%d'),
            'date_applied': body.get('date_applied', None),
            'priority': body.get('priority', 'medium'),
            'deadline': body.get('deadline', None),
        }
        try:
            insert_job(new_job)
        except Exception as e:
            print(f"DB insert error: {e}")
            return self.send_json(500, {'error': 'Failed to save'})
        return self.send_json(201, new_job)

    def do_PATCH(self):
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        path = parsed.path.rstrip('/')
        if path != '/api/jobs':
            return self.send_json(404, {'error': 'Not found'})
        job_id = qs.get('id', [''])[0]
        if not job_id:
            return self.send_json(400, {'error': 'Missing id'})
        length = int(self.headers.get('Content-Length', 0))
        body = json.loads(self.rfile.read(length).decode())
        allowed = ['status', 'notes', 'date_applied', 'priority', 'deadline', 'company',
                   'role', 'tier', 'category', 'url', 'careers_url', 'salary_range',
                   'location', 'description', 'source']
        fields = {f: body[f] for f in allowed if f in body}
        try:
            existing = get_job(job_id)
            if not existing:
                return self.send_json(404, {'error': 'Job not found'})
            if body.get('status') == 'applied' and not existing.get('date_applied') and not fields.get('date_applied'):
                fields['date_applied'] = datetime.now().strftime('%Y-%m-%d')
            update_job(job_id, fields)
            updated = get_job(job_id)
        except Exception as e:
            print(f"DB update error: {e}")
            return self.send_json(500, {'error': 'Failed to save'})
        return self.send_json(200, updated)

    def do_DELETE(self):
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        path = parsed.path.rstrip('/')
        if path != '/api/jobs':
            return self.send_json(404, {'error': 'Not found'})
        job_id = qs.get('id', [''])[0]
        if not job_id:
            return self.send_json(400, {'error': 'Missing id'})
        try:
            if not delete_job(job_id):
                return self.send_json(404, {'error': 'Job not found'})
        except Exception as e:
            print(f"DB delete error: {e}")
            return self.send_json(500, {'error': 'Failed to save'})
        return self.send_json(200, {'ok': True})
